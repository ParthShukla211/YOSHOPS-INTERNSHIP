#!/usr/bin/env python
# coding: utf-8

# # YOSHOPS INTERNSHIP TASK 2

# ## 1. Python program to Create a excel file

# In[5]:


# create a new XLSX workbook
wb = Workbook(FileFormatType.XLSX)
# insert value in the cells
wb.getWorksheets().get(0).getCells().get("A1").putValue("Hello World!")
# save workbook as .xlsx file
wb.save("workbook.xlsx")


# ## 2. Python program for Import data from an excel file 

# In[6]:


import pandas as pd

df = pd.read_excel ('file.xlxs')
print (df)


# ## 3. Python program for Format data in excel sheet 

# In[ ]:


import xlsxwriter

 # Create a workbook and add a worksheet.
 workbook = xlsxwriter.Workbook('Expenses02.xlsx')
 worksheet = workbook.add_worksheet()

 # Add a bold format to use to highlight cells.
 bold = workbook.add_format({'bold': True})

 # Add a number format for cells with money.
 money = workbook.add_format({'num_format': '$#,##0'})

 # Write some data headers.
 worksheet.write('A1', 'Item', bold)
 worksheet.write('B1', 'Cost', bold)

 # Some data we want to write to the worksheet.
 expenses = (
     ['Rent', 1000],
     ['Gas',   100],
     ['Food',  300],
     ['Gym',    50],
 )

 # Start from the first cell below the headers.
 row = 1
 col = 0

 # Iterate over the data and write it out row by row.
 for item, cost in (expenses):
     worksheet.write(row, col,     item)
     worksheet.write(row, col + 1, cost, money)
     row += 1

 # Write a total using a formula.
 worksheet.write(row, 0, 'Total',       bold)
 worksheet.write(row, 1, '=SUM(B2:B5)', money)

 workbook.close()
    
############################################

# Add a bold format to use to highlight cells.
bold = workbook.add_format({'bold': True})

# Add a number format for cells with money.
money = workbook.add_format({'num_format': '$#,##0'})


# ## 4.  Python program for Prepare excel charts 

# In[ ]:


# import openpyxl module
import openpyxl

# import BarChart class from openpyxl.chart sub_module
from openpyxl.chart import BarChart,Reference

# Call a Workbook() function of openpyxl
# to create a new blank Workbook object
wb = openpyxl.Workbook()

# Get workbook active sheet
# from the active attribute.
sheet = wb.active

# write o to 9 in 1st column of the active sheet
for i in range(10):
	sheet.append([i])

# create data for plotting
values = Reference(sheet, min_col = 1, min_row = 1,
						max_col = 1, max_row = 10)

# Create object of BarChart class
chart = BarChart()

# adding data to the Bar chart object
chart.add_data(values)

# set the title of the chart
chart.title = " BAR-CHART "

# set the title of the x-axis
chart.x_axis.title = " X_AXIS "

# set the title of the y-axis
chart.y_axis.title = " Y_AXIS "

# add chart to the sheet
# the top-left corner of a chart
# is anchored to cell E2 .
sheet.add_chart(chart, "E2")

# save the file
wb.save("barChart.xlsx")


# ## 5. Python program for Extract mobile no from PDF and MS word file and save into MS excel

# In[ ]:


import PyPDF2

pdffile = open("mkut-member-sample.pdf","rb")
pdf = PyPDF2.PdfFileReader(pdffile)
print(pdf.numPages)

pageobj = pdf.getPage(0)
print(pageobj.extractText())

pdffile.close()

